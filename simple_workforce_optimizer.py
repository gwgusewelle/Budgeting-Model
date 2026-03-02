import streamlit as st
import pandas as pd
import pulp

# ─── PAGE CONFIG ──────────────────────────────────────────────────────────────
# Set the browser tab title, icon, and use full-width layout
st.set_page_config(page_title="Workforce Optimizer", layout="wide")

# ─── SIDEBAR WIDTH ────────────────────────────────────────────────────────────
# Streamlit's default sidebar max-width is ~350px which clips longer labels.
# This CSS override bumps it to 420px and lets it resize freely up to that point.
st.markdown(
    """
    <style>
        [data-testid="stSidebar"] { min-width: 320px; max-width: 420px; }
        [data-testid="stSidebar"] .stFileUploader { width: 100%; }
    </style>
    """,
    unsafe_allow_html=True,
)


# ─── CONSTANTS ────────────────────────────────────────────────────────────────
# Emoji color indicators for each skill type used in display throughout the UI
SKILL_COLORS = {"A": "🟠", "B": "🔵", "C": "🟢", "D": "🟣", "E": "🔴"}

# Expected column names for each sheet in the uploaded Excel template.
# These are validated on upload to give the user clear feedback if something is missing.
REQUIRED_EMPLOYEE_COLS = {"employee", "capacity", "skills", "hourly_rate ($)"}
REQUIRED_PROJECT_COLS  = {"project", "reimbursable", "max_total", "budget ($)"}
REQUIRED_TASK_COLS     = {"project", "task_id", "task_type", "min_hours"}


# ─── BLANK TEMPLATE BUILDER ───────────────────────────────────────────────────
# Generates an in-memory Excel file with the three required sheets pre-populated
# with headers and one example row each, plus a Data Dictionary sheet.
# Returns raw bytes so Streamlit can serve it as a download without touching disk.
def build_blank_template() -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── color palette ──
    HEADER_FILL  = PatternFill("solid", start_color="1F4E79")   # dark navy
    EXAMPLE_FILL = PatternFill("solid", start_color="D9E1F2")   # light blue tint
    DICT_FILL    = PatternFill("solid", start_color="375623")    # dark green
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def style_header(cell, green=False):
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = DICT_FILL if green else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER

    def style_example(cell):
        cell.fill      = EXAMPLE_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = BORDER

    def style_note(cell):
        cell.font      = Font(italic=True, color="595959", size=9)
        cell.alignment = Alignment(horizontal="left")

    def set_col_widths(ws, widths):
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    # ── Employees sheet ──
    ws_emp = wb.active
    ws_emp.title = "Employees"
    emp_headers  = ["employee", "capacity", "skills", "hourly_rate ($)", "employee_type"]
    emp_example  = ["E001", 40, "A,B", 67.50, "Senior"]
    emp_example2 = ["E002", 30, "B,C,D", 55.00, "Mid-Level"]
    for c, h in enumerate(emp_headers, 1):
        style_header(ws_emp.cell(1, c, h))
    for c, v in enumerate(emp_example, 1):
        style_example(ws_emp.cell(2, c, v))
    for c, v in enumerate(emp_example2, 1):
        style_example(ws_emp.cell(3, c, v))
    # note row
    notes = ["Unique ID (e.g. E001)", "Max hrs/week (integer)",
             "Comma-separated A-E", "Hourly billing rate ($)", "Junior / Mid-Level / Senior"]
    for c, n in enumerate(notes, 1):
        style_note(ws_emp.cell(4, c, n))
    set_col_widths(ws_emp, [16, 14, 18, 18, 16])
    ws_emp.row_dimensions[1].height = 20

    # ── Projects sheet ──
    ws_proj = wb.create_sheet("Projects")
    proj_headers  = ["project", "reimbursable", "max_total", "budget ($)"]
    proj_example  = ["P001", True, 90, 12000]
    proj_example2 = ["P002", False, None, 8500]
    for c, h in enumerate(proj_headers, 1):
        style_header(ws_proj.cell(1, c, h))
    for c, v in enumerate(proj_example, 1):
        style_example(ws_proj.cell(2, c, v))
    for c, v in enumerate(proj_example2, 1):
        style_example(ws_proj.cell(3, c, v))
    notes_p = ["Unique ID (e.g. P001)", "TRUE = client-billable, FALSE = internal",
               "Max hours cap (leave blank if not reimbursable)", "Total dollar budget ($)"]
    for c, n in enumerate(notes_p, 1):
        style_note(ws_proj.cell(4, c, n))
    set_col_widths(ws_proj, [16, 16, 14, 14])
    ws_proj.row_dimensions[1].height = 20

    # ── Tasks sheet ──
    ws_task = wb.create_sheet("Tasks")
    task_headers  = ["project", "task_id", "task_type", "min_hours", "Urgency"]
    task_example  = ["P001", "T001", "A", 10, None]
    task_example2 = ["P001", "T002", "B", 14, None]
    for c, h in enumerate(task_headers, 1):
        style_header(ws_task.cell(1, c, h))
    for c, v in enumerate(task_example, 1):
        style_example(ws_task.cell(2, c, v))
    for c, v in enumerate(task_example2, 1):
        style_example(ws_task.cell(3, c, v))
    notes_t = ["Must match a project ID", "Unique task ID (e.g. T001)",
               "Skill type required (A / B / C / D / E)", "Minimum hours to complete", "Optional — leave blank"]
    for c, n in enumerate(notes_t, 1):
        style_note(ws_task.cell(4, c, n))
    set_col_widths(ws_task, [14, 14, 14, 14, 12])
    ws_task.row_dimensions[1].height = 20

    # ── Data Dictionary sheet ──
    ws_dict = wb.create_sheet("Data Dictionary")
    dict_headers = ["Sheet", "Column", "Type", "Description"]
    dict_rows = [
        ("Employees", "employee",       "ID",        "Unique employee identifier (e.g. E001)"),
        ("Employees", "capacity",       "Integer",   "Max weekly hours available"),
        ("Employees", "skills",         "String",    "Comma-separated skill types (A-E)"),
        ("Employees", "hourly_rate ($)","Float ($)",  "Hourly billing/wage rate"),
        ("Employees", "employee_type",  "Category",  "Junior (<25 h), Mid-Level (25-39 h), Senior (40+ h)"),
        ("Projects",  "project",        "ID",        "Unique project identifier (e.g. P001)"),
        ("Projects",  "reimbursable",   "Boolean",   "TRUE = client-billable; FALSE = internal/overhead"),
        ("Projects",  "max_total",      "Integer",   "Max hours budget for reimbursable projects (blank = no cap)"),
        ("Projects",  "budget ($)",     "Float ($)",  "Total dollar budget. LP uses this as a wage cost ceiling."),
        ("Tasks",     "project",        "ID",        "Project this task belongs to"),
        ("Tasks",     "task_id",        "ID",        "Unique task identifier (e.g. T001)"),
        ("Tasks",     "task_type",      "Category",  "Required skill type (A / B / C / D / E)"),
        ("Tasks",     "min_hours",      "Integer",   "Minimum hours required to complete task"),
        ("Tasks",     "Urgency",        "Optional",  "Not used by optimizer — reserved for future use"),
    ]
    for c, h in enumerate(dict_headers, 1):
        style_header(ws_dict.cell(1, c, h), green=True)
    for r, row in enumerate(dict_rows, 2):
        for c, v in enumerate(row, 1):
            cell        = ws_dict.cell(r, c, v)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
    set_col_widths(ws_dict, [14, 18, 12, 52])
    ws_dict.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── TEMPLATE VALIDATION ──────────────────────────────────────────────────────
# Checks that the uploaded file has the expected sheets and required columns.
# Returns (True, None) if valid or (False, error_message) if something is wrong.
def validate_template(xl: pd.ExcelFile) -> tuple[bool, str | None]:
    required_sheets = {"Employees", "Projects", "Tasks"}
    missing_sheets  = required_sheets - set(xl.sheet_names)
    if missing_sheets:
        return False, f"Missing sheet(s): {', '.join(missing_sheets)}. Expected: Employees, Projects, Tasks."

    emp_df  = pd.read_excel(xl, sheet_name="Employees", nrows=0)
    proj_df = pd.read_excel(xl, sheet_name="Projects",  nrows=0)
    task_df = pd.read_excel(xl, sheet_name="Tasks",     nrows=0)

    emp_cols  = set(emp_df.columns.str.strip())
    proj_cols = set(proj_df.columns.str.strip())
    task_cols = set(task_df.columns.str.strip())

    missing_emp  = REQUIRED_EMPLOYEE_COLS - emp_cols
    missing_proj = REQUIRED_PROJECT_COLS  - proj_cols
    missing_task = REQUIRED_TASK_COLS     - task_cols

    errors = []
    if missing_emp:
        errors.append(f"Employees sheet missing column(s): {', '.join(missing_emp)}")
    if missing_proj:
        errors.append(f"Projects sheet missing column(s): {', '.join(missing_proj)}")
    if missing_task:
        errors.append(f"Tasks sheet missing column(s): {', '.join(missing_task)}")

    if errors:
        return False, "\n".join(errors)
    return True, None


# ─── DATA LOADER ──────────────────────────────────────────────────────────────
# Reads the three core sheets from the uploaded Excel file and converts them
# into the list-of-dicts format the optimizer and UI expect.
# Strips whitespace from string fields and handles NaN gracefully.
def load_data(xl: pd.ExcelFile) -> tuple[list, list, list]:
    # --- Employees ---
    emp_df = pd.read_excel(xl, sheet_name="Employees")
    emp_df.columns = emp_df.columns.str.strip()
    employees = []
    for _, row in emp_df.iterrows():
        raw_skills = str(row["skills"]).strip()
        # Skills stored as comma-separated string (e.g. "A,B,C") → parse to list
        skills = [s.strip() for s in raw_skills.split(",") if s.strip()]
        employees.append({
            "id":       str(row["employee"]).strip(),
            "capacity": int(row["capacity"]),
            "skills":   skills,
            "rate":     float(row["hourly_rate ($)"]),   # hourly wage used in budget LP constraint
            "type":     str(row.get("employee_type", "")).strip(),
        })

    # --- Projects ---
    proj_df = pd.read_excel(xl, sheet_name="Projects")
    proj_df.columns = proj_df.columns.str.strip()
    projects = []
    for _, row in proj_df.iterrows():
        max_total = None if pd.isna(row["max_total"]) else int(row["max_total"])
        budget    = None if pd.isna(row["budget ($)"]) else float(row["budget ($)"])
        projects.append({
            "id":          str(row["project"]).strip(),
            "reimbursable": bool(row["reimbursable"]),
            "maxTotal":    max_total,   # max hours cap (reimbursable projects only)
            "budget":      budget,      # dollar budget cap used in LP wage constraint
        })

    # --- Tasks ---
    task_df = pd.read_excel(xl, sheet_name="Tasks")
    task_df.columns = task_df.columns.str.strip()
    tasks = []
    for _, row in task_df.iterrows():
        tasks.append({
            "project":  str(row["project"]).strip(),
            "id":       str(row["task_id"]).strip(),
            "type":     str(row["task_type"]).strip(),
            "minHours": int(row["min_hours"]),
        })

    return employees, projects, tasks


# ─── LINEAR PROGRAM OPTIMIZER ─────────────────────────────────────────────────
# Uses PuLP to formulate and solve a linear program that maximizes task coverage.
#
# Decision variables:
#   x[i][j]  = continuous hours employee i spends on task j  (>= 0)
#   s[j]     = unmet/slack hours for task j                  (>= 0)
#
# Objective: minimize sum(weight_j * s[j])
#   where weight_j is 2 for reimbursable-project tasks, 1 for non-reimbursable.
#   This ensures reimbursable tasks are prioritized by the optimizer.
#
# Constraints:
#   1. Employee capacity:  sum_j(x[i][j]) <= capacity[i]          for each employee i
#   2. Task demand:        sum_i(x[i][j]) + s[j] >= minHours[j]   for each task j
#   3. Skill match:        x[i][j] = 0  if employee i lacks the required skill for task j
#   4. Project hour cap:   sum_{j in p} sum_i(x[i][j]) <= maxTotal[p]  (reimbursable)
#   5. Project $ budget:   sum_{j in p} sum_i(x[i][j] * rate[i]) <= budget[p]  for each project p
#
# Returns a dict with:
#   asgn     - {task_id: {employee, hours, partial}} assignment map
#   load     - {employee_id: total hours assigned}
#   p_hours  - {project_id: total hours assigned}
#   p_cost   - {project_id: total wage cost assigned ($)}
#   status   - solver status string
def run_optimizer(employees, tasks, projects):
    proj_map = {p["id"]: p for p in projects}

    # Build index lookups for clean variable naming
    emp_ids  = [e["id"] for e in employees]
    task_ids = [t["id"] for t in tasks]
    emp_map  = {e["id"]: e for e in employees}
    task_map = {t["id"]: t for t in tasks}

    # Determine which (employee, task) pairs are valid based on skill match
    valid_pairs = [
        (eid, tid)
        for eid in emp_ids
        for tid in task_ids
        if task_map[tid]["type"] in emp_map[eid]["skills"]
    ]

    # Initialize the LP problem (minimization)
    prob = pulp.LpProblem("workforce_optimizer", pulp.LpMinimize)

    # --- Decision variables ---
    # x[(eid, tid)] = hours employee eid works on task tid
    x = {
        (eid, tid): pulp.LpVariable(f"x_{eid}_{tid}", lowBound=0)
        for (eid, tid) in valid_pairs
    }
    # s[tid] = unmet hours for task tid (slack variable)
    s = {
        tid: pulp.LpVariable(f"s_{tid}", lowBound=0)
        for tid in task_ids
    }

    # --- Objective: minimize weighted unmet hours ---
    # Reimbursable tasks carry weight 2 so the LP covers them first
    prob += pulp.lpSum(
        (2 if proj_map.get(task_map[tid]["project"], {}).get("reimbursable") else 1) * s[tid]
        for tid in task_ids
    )

    # --- Constraint 1: Employee capacity ---
    # Each employee cannot work more hours than their weekly capacity
    for eid in emp_ids:
        emp_tasks = [tid for (e, tid) in valid_pairs if e == eid]
        if emp_tasks:
            prob += pulp.lpSum(x[(eid, tid)] for tid in emp_tasks) <= emp_map[eid]["capacity"]

    # --- Constraint 2: Task demand (met hours + slack >= minHours) ---
    # Ensures every task is either covered by assigned employees or accounted for as unmet
    for tid in task_ids:
        assigned_vars = [x[(eid, tid)] for (eid, t) in valid_pairs if t == tid]
        if assigned_vars:
            prob += pulp.lpSum(assigned_vars) + s[tid] >= task_map[tid]["minHours"]
        else:
            # No eligible employee exists — slack must absorb the full demand
            prob += s[tid] >= task_map[tid]["minHours"]

    # --- Constraint 3: Project hour cap (reimbursable projects only) ---
    # Prevents total hours billed to a reimbursable project from exceeding maxTotal
    for p in projects:
        if p["reimbursable"] and p["maxTotal"] is not None:
            project_pairs = [(eid, tid) for (eid, tid) in valid_pairs if task_map[tid]["project"] == p["id"]]
            if project_pairs:
                prob += pulp.lpSum(x[(eid, tid)] for (eid, tid) in project_pairs) <= p["maxTotal"]

    # --- Constraint 4: Project dollar budget cap ---
    # Prevents the total wage cost (hours × rate) assigned to any project from exceeding its budget
    for p in projects:
        if p["budget"] is not None:
            project_pairs = [(eid, tid) for (eid, tid) in valid_pairs if task_map[tid]["project"] == p["id"]]
            if project_pairs:
                prob += pulp.lpSum(
                    x[(eid, tid)] * emp_map[eid]["rate"]
                    for (eid, tid) in project_pairs
                ) <= p["budget"]

    # --- Solve ---
    # Use CBC solver silently (msg=0 suppresses console output)
    prob.solve(pulp.PULP_CBC_CMD(msg=0))

    # --- Extract results ---
    # Build assignment map: for each task, find which employee has the most hours assigned
    asgn    = {}
    load    = {eid: 0.0 for eid in emp_ids}
    p_hours = {p["id"]: 0.0 for p in projects}
    p_cost  = {p["id"]: 0.0 for p in projects}

    # Aggregate hours assigned per task across all eligible employees
    task_assigned = {tid: {} for tid in task_ids}
    for (eid, tid) in valid_pairs:
        val = pulp.value(x[(eid, tid)])
        if val and val > 0.01:  # ignore near-zero floating point noise
            task_assigned[tid][eid] = val
            load[eid]                           += val
            p_hours[task_map[tid]["project"]]   += val
            p_cost[task_map[tid]["project"]]    += val * emp_map[eid]["rate"]

    # Determine the primary employee for each task (highest hours assigned)
    for tid in task_ids:
        assigned = task_assigned[tid]
        if not assigned:
            # No employee was assigned any hours to this task
            asgn[tid] = {"employee": None, "hours": 0, "reason": "no_skill_or_budget"}
        else:
            # Primary owner = employee with the most hours on this task
            primary_eid = max(assigned, key=assigned.get)
            hours       = sum(assigned.values())
            partial     = hours < task_map[tid]["minHours"] - 0.5  # flagged if short by >0.5 h
            asgn[tid]   = {
                "employee": primary_eid,
                "hours":    round(hours, 1),
                "partial":  partial,
                "all_assigned": assigned,   # full breakdown for multi-employee tasks
            }

    return {
        "asgn":    asgn,
        "load":    {k: round(v, 1) for k, v in load.items()},
        "p_hours": {k: round(v, 1) for k, v in p_hours.items()},
        "p_cost":  {k: round(v, 2) for k, v in p_cost.items()},
        "status":  pulp.LpStatus[prob.status],
    }


# ─── STYLE HELPER ─────────────────────────────────────────────────────────────
# Maps status strings to emoji-colored labels for consistent UI display
def status_badge(status):
    colors = {
        "OK":          "🟢", "Overloaded": "🟡", "Failed":     "🔴",
        "Fulfilled":   "🟢", "Partial":    "🟡", "Over Budget":"🔴",
        "critical":    "🔴", "tight":      "🟡", "healthy":    "🟢",
        "surplus":     "⚪", "Reassigned": "🟢",
    }
    return f"{colors.get(status, '')} {status}"


# ─── HEADER ───────────────────────────────────────────────────────────────────
st.markdown("## ⚡ Workforce Optimizer")
st.caption("Upload your budget data template to run the optimizer.")


# ─── FILE UPLOAD (SIDEBAR) ────────────────────────────────────────────────────
# The sidebar houses the upload widget so it stays accessible across all tabs.
# Users upload an Excel file matching the provided template structure.
with st.sidebar:
    st.header("📂 Data Import")
    st.markdown(
        "Upload an Excel file with three sheets: **Employees**, **Projects**, and **Tasks**."
    )

    # ── Downloadable blank template ──
    # Build the template fresh each session so the download is always current.
    # Wrapped in a try so a packaging issue never blocks the upload flow.
    try:
        template_bytes = build_blank_template()
        st.download_button(
            label="⬇ Download blank template",
            data=template_bytes,
            file_name="workforce_optimizer_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception:
        st.caption("Template download unavailable — openpyxl may not be installed.")

    st.divider()

    # ── Downloadable reference guide (.docx) ──
    try:
        with open("/sessions/gracious-friendly-keller/mnt/Optimizer/workforce_optimizer_guide.docx", "rb") as f:
            guide_bytes = f.read()
        st.download_button(
            label="⬇ Download reference guide (.docx)",
            data=guide_bytes,
            file_name="workforce_optimizer_guide.docx",
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            use_container_width=True,
        )
    except Exception:
        st.caption("Reference guide unavailable.")

    st.divider()

    uploaded_file = st.file_uploader(
        label="Choose Excel file (.xlsx)",
        type=["xlsx"],
        help="File must contain sheets: Employees, Projects, Tasks",
    )


# ─── GUARD: require upload before anything else runs ─────────────────────────
# If no file is uploaded, show instructions and stop rendering the rest of the app.
if uploaded_file is None:
    st.info("👈 Upload your Excel template in the sidebar to get started.")
    st.stop()


# ─── PARSE AND VALIDATE UPLOAD ────────────────────────────────────────────────
# Parse the uploaded bytes into a pandas ExcelFile, validate the structure,
# then load into the data structures the optimizer expects.
try:
    xl = pd.ExcelFile(uploaded_file)
except Exception as e:
    st.error(f"Could not read the file: {e}")
    st.stop()

valid, err_msg = validate_template(xl)
if not valid:
    st.error(f"**Invalid template:**\n\n{err_msg}")
    st.stop()

# Load data — wrapped in try/except to surface any parsing errors clearly
try:
    EMPLOYEES, PROJECTS, TASKS = load_data(xl)
except Exception as e:
    st.error(f"Error reading data from file: {e}")
    st.stop()

# Confirm successful load to the user
st.sidebar.success(
    f"Loaded: {len(EMPLOYEES)} employees · {len(PROJECTS)} projects · {len(TASKS)} tasks"
)


# ─── RUN OPTIMIZER ────────────────────────────────────────────────────────────
# Run the LP optimizer once on page load and cache results in session_state
# so switching tabs doesn't re-solve unnecessarily.
# The cache key is the file name + size so re-uploading a new file re-solves.
cache_key = f"{uploaded_file.name}_{uploaded_file.size}"
if st.session_state.get("_cache_key") != cache_key:
    with st.spinner("Running LP optimizer..."):
        opt_result = run_optimizer(EMPLOYEES, TASKS, PROJECTS)
    st.session_state["opt_result"]  = opt_result
    st.session_state["_cache_key"] = cache_key
else:
    opt_result = st.session_state["opt_result"]

# Extract optimization results and build convenience lookups
asgn       = opt_result["asgn"]
load       = opt_result["load"]
p_hours    = opt_result["p_hours"]
p_cost     = opt_result["p_cost"]

# Convenience lookups used across multiple tabs
proj_map     = {p["id"]: p for p in PROJECTS}
task_map     = {t["id"]: t for t in TASKS}

# Summary counts for the top-level metrics
total_cap  = sum(e["capacity"] for e in EMPLOYEES)
total_load = sum(load.values())
util       = round(total_load / total_cap * 100) if total_cap > 0 else 0
n_ok       = sum(1 for a in asgn.values() if a.get("employee") and not a.get("partial"))
n_fail     = sum(1 for a in asgn.values() if not a.get("employee"))
n_part     = sum(1 for a in asgn.values() if a.get("partial"))

# Build per-employee task list for the Employees tab
tasks_by_emp = {e["id"]: [] for e in EMPLOYEES}
for t in TASKS:
    a = asgn.get(t["id"], {})
    if a.get("employee"):
        tasks_by_emp[a["employee"]].append({**t, **a})


# ─── TOP METRICS ──────────────────────────────────────────────────────────────
# Three headline KPIs shown above the tabs for a quick status read
col1, col2, col3 = st.columns(3)
col1.metric("Utilization",   f"{util}%")
col2.metric("Tasks Covered", f"{n_ok} / {len(TASKS)}")
col3.metric("Unfilled",      f"{n_fail + n_part}")

# Surface the LP solver status in case the model is infeasible or unbounded
solver_status = opt_result.get("status", "Unknown")
if solver_status not in ("Optimal", "Not Solved"):
    st.warning(f"LP solver status: **{solver_status}** — results may be incomplete.")

st.divider()


# ─── TABS ─────────────────────────────────────────────────────────────────────
tab_labels = [
    "📊 Dashboard", "👥 Employees", "📁 Projects", "📋 Assignments",
]
tabs = st.tabs(tab_labels)


# ══════════════════════════════════════════════════════════════════ DASHBOARD
with tabs[0]:
    # Summary KPI row
    c1, c2, c3, c4 = st.columns(4)
    reimb_count = sum(1 for p in PROJECTS if p["reimbursable"])
    c1.metric("Total Employees",     len(EMPLOYEES), f"{total_cap} total h/week")
    c2.metric("Active Projects",     len(PROJECTS),
              f"{reimb_count} reimbursable · {len(PROJECTS)-reimb_count} standard")
    c3.metric("Tasks Assigned",      f"{n_ok} / {len(TASKS)}",
              f"{n_fail + n_part} need attention" if n_fail + n_part > 0 else "All tasks covered")
    c4.metric("Workforce Utilization", f"{util}%",
              f"{round(total_load, 1)} of {total_cap} h used")

    # Employee utilization table with progress bar
    st.subheader("Employee Utilization")
    emp_rows = []
    for e in EMPLOYEES:
        l = load.get(e["id"], 0)
        p = round(l / e["capacity"] * 100) if e["capacity"] > 0 else 0
        emp_rows.append({
            "Employee":      e["id"],
            "Type":          e.get("type", ""),
            "Rate ($/h)":    f"${e['rate']:.2f}",
            "Skills":        " ".join(SKILL_COLORS.get(s, s) for s in e["skills"]),
            "Assigned (h)":  round(l, 1),
            "Capacity (h)":  e["capacity"],
            "Utilization %": p,
        })
    df_emp = pd.DataFrame(emp_rows)
    st.dataframe(
        df_emp, use_container_width=True, hide_index=True,
        column_config={"Utilization %": st.column_config.ProgressColumn(
            "Utilization %", min_value=0, max_value=100, format="%d%%"
        )},
    )

    # Skill demand vs supply table
    st.subheader("Skill Demand vs Supply")
    skill_types = sorted(set(t["type"] for t in TASKS))
    skill_rows = []
    for sk in skill_types:
        dem    = sum(t["minHours"] for t in TASKS if t["type"] == sk)
        sup    = sum(e["capacity"] for e in EMPLOYEES if sk in e["skills"])
        ratio  = dem / sup if sup > 0 else 99
        status = ("surplus" if ratio <= 0.6 else "healthy" if ratio <= 1.0
                  else "tight" if ratio <= 1.3 else "critical")
        n_emp  = len([e for e in EMPLOYEES if sk in e["skills"]])
        n_task = len([t for t in TASKS if t["type"] == sk])
        skill_rows.append({
            "Skill":       f"{SKILL_COLORS.get(sk, sk)} {sk}",
            "Tasks":       n_task, "Employees": n_emp,
            "Demand (h)":  dem, "Supply (h)": sup,
            "Coverage %":  round(min(sup / dem, 1) * 100) if dem > 0 else 100,
            "Status":      status.upper(),
        })
    df_skills = pd.DataFrame(skill_rows)
    st.dataframe(
        df_skills, use_container_width=True, hide_index=True,
        column_config={"Coverage %": st.column_config.ProgressColumn(
            "Coverage %", min_value=0, max_value=100, format="%d%%"
        )},
    )

    # Project summary table including dollar budget tracking
    st.subheader("Project Summary")
    proj_rows = []
    for p in PROJECTS:
        dem     = sum(t["minHours"] for t in TASKS if t["project"] == p["id"])
        asn     = p_hours.get(p["id"], 0)
        cost    = p_cost.get(p["id"], 0)
        cov_pct = round(asn / dem * 100) if dem > 0 else 100
        ov_hr   = p["maxTotal"] and asn > p["maxTotal"]
        ov_bud  = p["budget"]   and cost > p["budget"]
        ok      = asn >= dem
        status  = "Over Budget" if (ov_hr or ov_bud) else ("Fulfilled" if ok else "Partial")
        proj_rows.append({
            "Project":       p["id"],
            "Type":          "Reimbursable" if p["reimbursable"] else "Non-Reimb.",
            "Demand (h)":    dem,
            "Assigned (h)":  round(asn, 1),
            "Hour Cap":      f"{p['maxTotal']} h" if p["maxTotal"] else "—",
            "Wage Cost ($)": f"${cost:,.0f}",
            "Budget ($)":    f"${p['budget']:,.0f}" if p["budget"] else "—",
            "Coverage %":    cov_pct,
            "Status":        status_badge(status),
        })
    st.dataframe(
        pd.DataFrame(proj_rows), use_container_width=True, hide_index=True,
        column_config={"Coverage %": st.column_config.ProgressColumn(
            "Coverage %", min_value=0, max_value=100, format="%d%%"
        )},
    )


# ══════════════════════════════════════════════════════════════════ EMPLOYEES
with tabs[1]:
    st.subheader("Employee Roster")
    for e in EMPLOYEES:
        l          = load.get(e["id"], 0)
        p          = round(l / e["capacity"] * 100) if e["capacity"] > 0 else 0
        skill_icons = " ".join(SKILL_COLORS.get(s, s) + s for s in e["skills"])
        label      = (f"**{e['id']}** [{e.get('type','')}] — {skill_icons} — "
                      f"${e['rate']:.2f}/h — {p}% utilized ({round(l,1)}/{e['capacity']} h)")
        with st.expander(label):
            tasks = tasks_by_emp.get(e["id"], [])
            if not tasks:
                st.write("No tasks assigned.")
            else:
                rows = [{
                    "Task":    t["id"],
                    "Project": t["project"],
                    "Skill":   f"{SKILL_COLORS.get(t['type'], t['type'])} {t['type']}",
                    "Min Hrs": t["minHours"],
                    "Assigned Hrs": round(t.get("hours", t["minHours"]), 1),
                    "Status":  "Partial" if t.get("partial") else "OK",
                } for t in tasks]
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════ PROJECTS
with tabs[2]:
    st.subheader("Project Overview")
    for p in PROJECTS:
        pt    = [t for t in TASKS if t["project"] == p["id"]]
        dem   = sum(t["minHours"] for t in pt)
        asn   = p_hours.get(p["id"], 0)
        cost  = p_cost.get(p["id"], 0)
        ov_hr = p["maxTotal"] and asn > p["maxTotal"]
        ov_bd = p["budget"]   and cost > p["budget"]
        ok    = asn >= dem
        label = (
            "🟢 Fulfilled"   if ok and not ov_hr and not ov_bd
            else "🔴 Over Budget" if (ov_hr or ov_bd)
            else "🟡 Partial"
        )
        ptype = "Reimbursable" if p["reimbursable"] else "Non-Reimbursable"
        header = (f"**{p['id']}** — {ptype} — {label} "
                  f"({round(asn,1)}/{dem} h · ${cost:,.0f} cost)")
        with st.expander(header):
            # Show hour budget progress bar for reimbursable projects
            if p["maxTotal"]:
                st.progress(
                    min(asn / p["maxTotal"], 1.0),
                    text=f"Hour budget: {round(asn,1)} / {p['maxTotal']} h",
                )
            # Show dollar budget progress bar if a budget is set
            if p["budget"]:
                st.progress(
                    min(cost / p["budget"], 1.0),
                    text=f"Wage budget: ${cost:,.0f} / ${p['budget']:,.0f}",
                )

            # ── Cost Breakdown section ──
            st.markdown("##### 💰 Cost Breakdown")
            cost_rows = []
            project_tasks = [t for t in TASKS if t["project"] == p["id"]]
            for t in project_tasks:
                a = asgn.get(t["id"], {})
                emp_id = a.get("employee")
                if emp_id:
                    emp = next((e for e in EMPLOYEES if e["id"] == emp_id), None)
                    if emp:
                        hrs = a.get("hours", 0)
                        task_cost = round(hrs * emp["rate"], 2)
                        cost_rows.append({
                            "Employee": emp_id,
                            "Tasks Assigned": 1,
                            "Hours": round(hrs, 1),
                            "Rate ($/h)": f"${emp['rate']:.2f}",
                            "Cost ($)": f"${task_cost:,.2f}",
                        })
            
            # Aggregate by employee
            if cost_rows:
                cost_df = pd.DataFrame(cost_rows)
                agg_rows = []
                for emp_id in cost_df["Employee"].unique():
                    emp_data = cost_df[cost_df["Employee"] == emp_id]
                    total_tasks = len(emp_data)
                    total_hours = emp_data["Hours"].sum()
                    emp_rate = emp_data["Rate ($/h)"].iloc[0]
                    emp_cost = round(total_hours * float(emp_rate.strip("$")), 2)
                    agg_rows.append({
                        "Employee": emp_id,
                        "Tasks Assigned": total_tasks,
                        "Hours": round(total_hours, 1),
                        "Rate ($/h)": emp_rate,
                        "Cost ($)": f"${emp_cost:,.2f}",
                    })
                
                # Add totals row
                agg_rows.append({
                    "Employee": "TOTAL",
                    "Tasks Assigned": sum(r["Tasks Assigned"] for r in agg_rows),
                    "Hours": round(sum(r["Hours"] for r in agg_rows), 1),
                    "Rate ($/h)": "—",
                    "Cost ($)": f"${cost:,.2f}",
                })
                
                st.dataframe(pd.DataFrame(agg_rows), use_container_width=True, hide_index=True)
            else:
                st.caption("No tasks assigned to this project.")

            # Task assignment details
            st.markdown("##### Task Details")
            rows = []
            for t in pt:
                a = asgn.get(t["id"], {})
                rows.append({
                    "Task":        t["id"],
                    "Skill":       f"{SKILL_COLORS.get(t['type'], t['type'])} {t['type']}",
                    "Min Hrs":     t["minHours"],
                    "Assigned Hrs":round(a.get("hours", 0), 1),
                    "Assigned To": a.get("employee") or "—",
                    "Status": (
                        "🟡 Partial" if a.get("partial")
                        else "🟢 OK"   if a.get("employee")
                        else "🔴 Unassigned"
                    ),
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════ ASSIGNMENTS
with tabs[3]:
    st.subheader("Full Task Assignment Matrix")
    rows = []
    for t in TASKS:
        a     = asgn.get(t["id"], {})
        emp_id = a.get("employee")
        
        # Calculate cost for this task
        cost_val = "—"
        if emp_id:
            emp = next((e for e in EMPLOYEES if e["id"] == emp_id), None)
            if emp:
                hrs = a.get("hours", 0)
                cost_val = f"${round(hrs * emp['rate'], 2):,.2f}"
        
        s_str = "Failed" if not a.get("employee") else ("Overloaded" if a.get("partial") else "OK")
        reimb = "R" if proj_map.get(t["project"], {}).get("reimbursable") else "N"
        rows.append({
            "Task":        t["id"],
            "Project":     f"{t['project']} ({reimb})",
            "Skill":       f"{SKILL_COLORS.get(t['type'], t['type'])} {t['type']}",
            "Min Hours":   t["minHours"],
            "Assigned Hrs":round(a.get("hours", 0), 1),
            "Assigned To": a.get("employee") or "—",
            "Cost ($)":    cost_val,
            "Status":      status_badge(s_str),
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
