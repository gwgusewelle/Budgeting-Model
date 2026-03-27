import io
import streamlit as st
import pandas as pd
import pulp
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Workforce Optimizer", layout="wide")

st.markdown(
    "<style>[data-testid='stSidebar']{min-width:320px;max-width:420px}</style>",
    unsafe_allow_html=True,
)

EXPECTED_SCHEMA = {
    "Employees": {"employee", "capacity", "skills", "hourly_rate ($)"},
    "Projects":  {"project", "reimbursable", "max_total", "budget ($)"},
    "Tasks":     {"project", "task_id", "task_type", "min_hours"},
}


def create_template_workbook() -> bytes:
    wb = Workbook()

    def populate_sheet(ws, column_headers, sample_data, column_notes, col_widths):
        for col_idx, header in enumerate(column_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row_idx, data_row in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(data_row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        note_row = len(sample_data) + 2
        for col_idx, note in enumerate(column_notes, start=1):
            cell = ws.cell(row=note_row, column=col_idx, value=note)
            cell.font = Font(italic=True, size=9)
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    employees_ws = wb.active
    employees_ws.title = "Employees"
    populate_sheet(
        employees_ws,
        ["employee", "capacity", "skills", "hourly_rate ($)", "employee_type"],
        [["E001", 40, "A,B", 67.50, "Senior"], ["E002", 30, "B,C,D", 55.00, "Mid-Level"]],
        ["Unique ID", "Max hrs/week", "Comma-sep A-E", "Hourly rate ($)", "Junior/Mid-Level/Senior"],
        [16, 14, 18, 18, 16],
    )

    projects_ws = wb.create_sheet("Projects")
    populate_sheet(
        projects_ws,
        ["project", "reimbursable", "max_total", "budget ($)"],
        [["P001", True, 90, 12000], ["P002", False, None, 8500]],
        ["Unique ID", "TRUE=billable FALSE=internal", "Max hours (blank if non-reimb.)", "Dollar budget"],
        [16, 16, 14, 14],
    )

    tasks_ws = wb.create_sheet("Tasks")
    populate_sheet(
        tasks_ws,
        ["project", "task_id", "task_type", "min_hours", "Urgency"],
        [["P001", "T001", "A", 10, None], ["P001", "T002", "B", 14, None]],
        ["Match a project ID", "Unique task ID", "Skill A/B/C/D/E", "Min hours", "Optional"],
        [14, 14, 14, 14, 12],
    )

    dict_ws = wb.create_sheet("Data Dictionary")
    dictionary_headers = ["Sheet", "Column", "Type", "Description"]
    dictionary_entries = [
        ("Employees", "employee",        "ID",       "Unique employee identifier"),
        ("Employees", "capacity",        "Integer",  "Max weekly hours available"),
        ("Employees", "skills",          "String",   "Comma-separated skill types A-E"),
        ("Employees", "hourly_rate ($)", "Float",    "Hourly billing/wage rate"),
        ("Employees", "employee_type",   "Category", "Junior / Mid-Level / Senior"),
        ("Projects",  "project",         "ID",       "Unique project identifier"),
        ("Projects",  "reimbursable",    "Boolean",  "TRUE = client-billable"),
        ("Projects",  "max_total",       "Integer",  "Max hours cap (blank if non-reimbursable)"),
        ("Projects",  "budget ($)",      "Float",    "Dollar budget ceiling for wage costs"),
        ("Tasks",     "project",         "ID",       "Must match a project ID"),
        ("Tasks",     "task_id",         "ID",       "Unique task identifier"),
        ("Tasks",     "task_type",       "Category", "Required skill type (A-E)"),
        ("Tasks",     "min_hours",       "Integer",  "Minimum hours to complete task"),
        ("Tasks",     "Urgency",         "Optional", "Not used by optimizer"),
    ]
    for col_idx, header in enumerate(dictionary_headers, start=1):
        dict_ws.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
    for row_idx, entry in enumerate(dictionary_entries, start=2):
        for col_idx, value in enumerate(entry, start=1):
            dict_ws.cell(row=row_idx, column=col_idx, value=value)
    for col_idx, width in enumerate([14, 18, 12, 50], start=1):
        dict_ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def check_template_validity(excel_file):
    required_sheets = set(EXPECTED_SCHEMA.keys())
    available_sheets = set(excel_file.sheet_names)
    missing_sheets = required_sheets - available_sheets

    if missing_sheets:
        return False, f"Missing sheet(s): {', '.join(sorted(missing_sheets))}"

    validation_errors = []
    for sheet_name, required_columns in EXPECTED_SCHEMA.items():
        df_headers = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=0)
        actual_columns = set(df_headers.columns.str.strip())
        missing_columns = required_columns - actual_columns
        if missing_columns:
            validation_errors.append(f"{sheet_name} missing: {', '.join(sorted(missing_columns))}")

    if validation_errors:
        return False, "\n".join(validation_errors)
    return True, None


def parse_uploaded_data(excel_file):
    def clean_column_names(dataframe):
        dataframe.columns = dataframe.columns.str.strip()
        return dataframe

    emp_df = clean_column_names(pd.read_excel(excel_file, sheet_name="Employees"))
    employee_list = []
    for _, row in emp_df.iterrows():
        skill_string = str(row["skills"])
        parsed_skills = [s.strip() for s in skill_string.split(",") if s.strip()]
        employee_list.append({
            "id":       str(row["employee"]).strip(),
            "capacity": int(row["capacity"]),
            "skills":   parsed_skills,
            "rate":     float(row["hourly_rate ($)"]),
            "type":     str(row.get("employee_type", "")).strip(),
        })

    proj_df = clean_column_names(pd.read_excel(excel_file, sheet_name="Projects"))
    project_list = []
    for _, row in proj_df.iterrows():
        max_hours = None if pd.isna(row["max_total"]) else int(row["max_total"])
        budget_val = None if pd.isna(row["budget ($)"]) else float(row["budget ($)"])
        project_list.append({
            "id":           str(row["project"]).strip(),
            "reimbursable": bool(row["reimbursable"]),
            "maxTotal":     max_hours,
            "budget":       budget_val,
        })

    task_df = clean_column_names(pd.read_excel(excel_file, sheet_name="Tasks"))
    task_list = []
    for _, row in task_df.iterrows():
        task_list.append({
            "project":  str(row["project"]).strip(),
            "id":       str(row["task_id"]).strip(),
            "type":     str(row["task_type"]).strip(),
            "minHours": int(row["min_hours"]),
        })

    return employee_list, project_list, task_list


def solve_allocation(employees, tasks, projects):
    emp_lookup  = {e["id"]: e for e in employees}
    task_lookup = {t["id"]: t for t in tasks}
    proj_lookup = {p["id"]: p for p in projects}

    valid_pairs = [
        (emp["id"], task["id"])
        for emp in employees
        for task in tasks
        if task["type"] in emp["skills"]
    ]

    problem = pulp.LpProblem("workforce_allocation", pulp.LpMinimize)

    hours_var = {
        (emp_id, task_id): pulp.LpVariable(f"x_{emp_id}_{task_id}", lowBound=0)
        for emp_id, task_id in valid_pairs
    }
    slack_var = {
        task["id"]: pulp.LpVariable(f"slack_{task['id']}", lowBound=0)
        for task in tasks
    }

    problem += pulp.lpSum(
        (2 if proj_lookup.get(task_lookup[tid]["project"], {}).get("reimbursable") else 1) * slack_var[tid]
        for tid in slack_var
    )

    for emp in employees:
        assigned_tasks = [tid for eid, tid in valid_pairs if eid == emp["id"]]
        if assigned_tasks:
            problem += (
                pulp.lpSum(hours_var[(emp["id"], tid)] for tid in assigned_tasks)
                <= emp["capacity"]
            )

    for task in tasks:
        task_assignments = [hours_var[(eid, task["id"])] for eid, tid in valid_pairs if tid == task["id"]]
        if task_assignments:
            problem += pulp.lpSum(task_assignments) + slack_var[task["id"]] >= task["minHours"]
        else:
            problem += slack_var[task["id"]] >= task["minHours"]

    for proj in projects:
        if proj["reimbursable"] and proj["maxTotal"] is not None:
            project_pairs = [(eid, tid) for eid, tid in valid_pairs if task_lookup[tid]["project"] == proj["id"]]
            if project_pairs:
                problem += pulp.lpSum(hours_var[pair] for pair in project_pairs) <= proj["maxTotal"]

    for proj in projects:
        if proj["budget"] is not None:
            project_pairs = [(eid, tid) for eid, tid in valid_pairs if task_lookup[tid]["project"] == proj["id"]]
            if project_pairs:
                problem += (
                    pulp.lpSum(hours_var[(eid, tid)] * emp_lookup[eid]["rate"] for eid, tid in project_pairs)
                    <= proj["budget"]
                )

    problem.solve(pulp.PULP_CBC_CMD(msg=0))

    employee_load   = {emp["id"]: 0.0 for emp in employees}
    project_hours   = {proj["id"]: 0.0 for proj in projects}
    project_costs   = {proj["id"]: 0.0 for proj in projects}
    task_assignments = {task["id"]: {} for task in tasks}

    for emp_id, task_id in valid_pairs:
        val = pulp.value(hours_var[(emp_id, task_id)])
        if val is not None and val > 0.01:
            task_assignments[task_id][emp_id] = val
            employee_load[emp_id] += val
            parent_project = task_lookup[task_id]["project"]
            project_hours[parent_project] += val
            project_costs[parent_project] += val * emp_lookup[emp_id]["rate"]

    assignment_summary = {}
    for task in tasks:
        assigned_employees = task_assignments[task["id"]]
        if not assigned_employees:
            assignment_summary[task["id"]] = {"employee": None, "hours": 0}
        else:
            primary_employee = max(assigned_employees, key=assigned_employees.get)
            total_hours = sum(assigned_employees.values())
            assignment_summary[task["id"]] = {
                "employee": primary_employee,
                "hours":    round(total_hours, 1),
                "partial":  total_hours < task["minHours"] - 0.5,
            }

    return {
        "asgn":    assignment_summary,
        "load":    {k: round(v, 1) for k, v in employee_load.items()},
        "p_hours": {k: round(v, 1) for k, v in project_hours.items()},
        "p_cost":  {k: round(v, 2) for k, v in project_costs.items()},
        "status":  pulp.LpStatus[problem.status],
    }


st.markdown("## Workforce Optimizer")
st.caption("Upload your data template in the sidebar to run the optimizer.")

with st.sidebar:
    st.header("Data Import")
    st.markdown("Upload an Excel file with three sheets: **Employees**, **Projects**, and **Tasks**.")

    try:
        st.download_button(
            "Download blank template",
            data=create_template_workbook(),
            file_name="workforce_optimizer_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception:
        st.caption("Template download unavailable.")

    try:
        with open("Capstone Budget Data.xlsx", "rb") as f:
            st.download_button(
                "Download example dataset",
                data=f.read(),
                file_name="Capstone Budget Data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    except Exception:
        st.caption("Example dataset unavailable.")

    st.divider()
    uploaded_file = st.file_uploader("Choose Excel file (.xlsx)", type=["xlsx"])

if uploaded_file is None:
    st.info("Upload your Excel template in the sidebar to get started.")
    st.stop()

try:
    excel_data = pd.ExcelFile(uploaded_file)
except Exception as e:
    st.error(f"Could not read file: {e}")
    st.stop()

is_valid, error_message = check_template_validity(excel_data)
if not is_valid:
    st.error(f"Invalid template:\n\n{error_message}")
    st.stop()

try:
    EMPLOYEES, PROJECTS, TASKS = parse_uploaded_data(excel_data)
except Exception as e:
    st.error(f"Error loading data: {e}")
    st.stop()

st.sidebar.success(f"{len(EMPLOYEES)} employees · {len(PROJECTS)} projects · {len(TASKS)} tasks")

cache_key = f"{uploaded_file.name}_{uploaded_file.size}"
if st.session_state.get("_cache_key") != cache_key:
    with st.spinner("Running optimizer..."):
        st.session_state["opt_result"] = solve_allocation(EMPLOYEES, TASKS, PROJECTS)
        st.session_state["_cache_key"] = cache_key

results     = st.session_state["opt_result"]
assignments = results["asgn"]
emp_load    = results["load"]
proj_hours  = results["p_hours"]
proj_costs  = results["p_cost"]

proj_map = {p["id"]: p for p in PROJECTS}
emp_map  = {e["id"]: e for e in EMPLOYEES}

total_capacity = sum(e["capacity"] for e in EMPLOYEES)
total_assigned = round(sum(emp_load.values()), 1)
utilization_pct = round(total_assigned / total_capacity * 100) if total_capacity else 0
tasks_fully_covered = sum(1 for a in assignments.values() if a.get("employee") and not a.get("partial"))
tasks_unassigned = sum(1 for a in assignments.values() if not a.get("employee"))
tasks_partial = sum(1 for a in assignments.values() if a.get("partial"))

employee_tasks = {e["id"]: [] for e in EMPLOYEES}
for task in TASKS:
    a = assignments.get(task["id"], {})
    if a.get("employee"):
        employee_tasks[a["employee"]].append({**task, **a})

col1, col2, col3 = st.columns(3)
col1.metric("Utilization", f"{utilization_pct}%")
col2.metric("Tasks Covered", f"{tasks_fully_covered} / {len(TASKS)}")
col3.metric("Unfilled", tasks_unassigned + tasks_partial)

if results["status"] not in ("Optimal", "Not Solved"):
    st.warning(f"Solver status: {results['status']} — results may be incomplete.")

st.divider()

tab_dashboard, tab_employees, tab_projects, tab_assignments = st.tabs(
    ["Dashboard", "Employees", "Projects", "Assignments"]
)


with tab_dashboard:
    num_reimbursable = sum(1 for p in PROJECTS if p["reimbursable"])
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Employees", len(EMPLOYEES), f"{total_capacity} h/week capacity")
    c2.metric("Projects", len(PROJECTS), f"{num_reimbursable} reimbursable")
    c3.metric(
        "Tasks",
        f"{tasks_fully_covered} / {len(TASKS)}",
        f"{tasks_unassigned + tasks_partial} need attention" if (tasks_unassigned + tasks_partial) else "All covered",
    )
    c4.metric("Utilization", f"{utilization_pct}%", f"{total_assigned} of {total_capacity} h used")

    st.subheader("Employee Utilization")
    utilization_data = []
    for emp in EMPLOYEES:
        emp_hours = emp_load.get(emp["id"], 0)
        emp_util = round(emp_hours / emp["capacity"] * 100) if emp["capacity"] else 0
        utilization_data.append({
            "Employee":      emp["id"],
            "Type":          emp["type"],
            "Rate ($/h)":    f"${emp['rate']:.2f}",
            "Skills":        ", ".join(emp["skills"]),
            "Assigned (h)":  emp_hours,
            "Capacity (h)":  emp["capacity"],
            "Utilization %": emp_util,
        })
    st.dataframe(
        pd.DataFrame(utilization_data),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Utilization %": st.column_config.ProgressColumn(
                "Utilization %", min_value=0, max_value=100, format="%d%%"
            )
        },
    )

    st.subheader("Skill Demand vs Supply")
    skill_analysis = []
    all_skill_types = sorted(set(t["type"] for t in TASKS))
    for skill in all_skill_types:
        demand = sum(t["minHours"] for t in TASKS if t["type"] == skill)
        supply = sum(e["capacity"] for e in EMPLOYEES if skill in e["skills"])
        demand_ratio = demand / supply if supply else 99
        coverage = round(min(supply / demand, 1) * 100) if demand else 100

        if demand_ratio > 1.3:
            status = "CRITICAL"
        elif demand_ratio > 1.0:
            status = "TIGHT"
        elif demand_ratio > 0.6:
            status = "HEALTHY"
        else:
            status = "SURPLUS"

        skill_analysis.append({
            "Skill":      skill,
            "Employees":  sum(1 for e in EMPLOYEES if skill in e["skills"]),
            "Tasks":      sum(1 for t in TASKS if t["type"] == skill),
            "Demand (h)": demand,
            "Supply (h)": supply,
            "Coverage %": coverage,
            "Status":     status,
        })

    st.dataframe(
        pd.DataFrame(skill_analysis),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Coverage %": st.column_config.ProgressColumn(
                "Coverage %", min_value=0, max_value=100, format="%d%%"
            )
        },
    )

    st.subheader("Project Summary")
    project_summary = []
    for proj in PROJECTS:
        demand = sum(t["minHours"] for t in TASKS if t["project"] == proj["id"])
        assigned = proj_hours.get(proj["id"], 0)
        cost = proj_costs.get(proj["id"], 0)
        over_budget = (
            (proj["maxTotal"] is not None and assigned > proj["maxTotal"])
            or (proj["budget"] is not None and cost > proj["budget"])
        )

        if over_budget:
            proj_status = "Over Budget"
        elif assigned >= demand:
            proj_status = "Fulfilled"
        else:
            proj_status = "Partial"

        project_summary.append({
            "Project":       proj["id"],
            "Type":          "Reimbursable" if proj["reimbursable"] else "Non-Reimb.",
            "Demand (h)":    demand,
            "Assigned (h)":  round(assigned, 1),
            "Hour Cap":      f"{proj['maxTotal']} h" if proj["maxTotal"] else "\u2014",
            "Wage Cost ($)": f"${cost:,.0f}",
            "Budget ($)":    f"${proj['budget']:,.0f}" if proj["budget"] else "\u2014",
            "Coverage %":    round(assigned / demand * 100) if demand else 100,
            "Status":        proj_status,
        })

    st.dataframe(
        pd.DataFrame(project_summary),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Coverage %": st.column_config.ProgressColumn(
                "Coverage %", min_value=0, max_value=100, format="%d%%"
            )
        },
    )


with tab_employees:
    st.subheader("Employee Roster")
    for emp in EMPLOYEES:
        emp_hours = emp_load.get(emp["id"], 0)
        emp_util = round(emp_hours / emp["capacity"] * 100) if emp["capacity"] else 0
        expander_label = (
            f"**{emp['id']}** [{emp['type']}]  |  Skills: {', '.join(emp['skills'])}"
            f"  |  ${emp['rate']:.2f}/h  |  {emp_util}% utilized ({emp_hours}/{emp['capacity']} h)"
        )
        with st.expander(expander_label):
            emp_task_list = employee_tasks.get(emp["id"], [])
            if not emp_task_list:
                st.caption("No tasks assigned.")
            else:
                st.dataframe(
                    pd.DataFrame([{
                        "Task":         t["id"],
                        "Project":      t["project"],
                        "Skill":        t["type"],
                        "Min Hrs":      t["minHours"],
                        "Assigned Hrs": round(t.get("hours", t["minHours"]), 1),
                        "Status":       "Partial" if t.get("partial") else "OK",
                    } for t in emp_task_list]),
                    use_container_width=True,
                    hide_index=True,
                )


with tab_projects:
    st.subheader("Project Overview")
    for proj in PROJECTS:
        proj_tasks = [t for t in TASKS if t["project"] == proj["id"]]
        demand = sum(t["minHours"] for t in proj_tasks)
        assigned = proj_hours.get(proj["id"], 0)
        cost = proj_costs.get(proj["id"], 0)
        over_budget = (
            (proj["maxTotal"] is not None and assigned > proj["maxTotal"])
            or (proj["budget"] is not None and cost > proj["budget"])
        )

        if over_budget:
            status_text = "Over Budget"
        elif assigned >= demand:
            status_text = "Fulfilled"
        else:
            status_text = "Partial"

        header_text = (
            f"**{proj['id']}**  |  {'Reimbursable' if proj['reimbursable'] else 'Non-Reimbursable'}"
            f"  |  {status_text}  |  {round(assigned, 1)}/{demand} h  |  ${cost:,.0f}"
        )

        with st.expander(header_text):
            if proj["maxTotal"]:
                st.progress(
                    min(assigned / proj["maxTotal"], 1.0),
                    text=f"Hour budget: {round(assigned, 1)} / {proj['maxTotal']} h",
                )
            if proj["budget"]:
                st.progress(
                    min(cost / proj["budget"], 1.0),
                    text=f"Wage budget: ${cost:,.0f} / ${proj['budget']:,.0f}",
                )

            st.markdown("##### Cost Breakdown")
            cost_breakdown = {}
            for task in proj_tasks:
                a = assignments.get(task["id"], {})
                assigned_emp = a.get("employee")
                if assigned_emp:
                    if assigned_emp not in cost_breakdown:
                        cost_breakdown[assigned_emp] = {"tasks": 0, "hours": 0.0}
                    cost_breakdown[assigned_emp]["tasks"] += 1
                    cost_breakdown[assigned_emp]["hours"] += a.get("hours", 0)

            if cost_breakdown:
                breakdown_rows = [{
                    "Employee":   eid,
                    "Tasks":      info["tasks"],
                    "Hours":      round(info["hours"], 1),
                    "Rate ($/h)": f"${emp_map[eid]['rate']:.2f}",
                    "Cost ($)":   f"${info['hours'] * emp_map[eid]['rate']:,.2f}",
                } for eid, info in cost_breakdown.items()]

                breakdown_rows.append({
                    "Employee":   "TOTAL",
                    "Tasks":      sum(r["Tasks"] for r in breakdown_rows),
                    "Hours":      round(sum(r["Hours"] for r in breakdown_rows), 1),
                    "Rate ($/h)": "\u2014",
                    "Cost ($)":   f"${cost:,.2f}",
                })
                st.dataframe(pd.DataFrame(breakdown_rows), use_container_width=True, hide_index=True)
            else:
                st.caption("No tasks assigned to this project.")

            st.markdown("##### Task Details")
            task_detail_rows = []
            for task in proj_tasks:
                a = assignments.get(task["id"], {})
                if a.get("partial"):
                    task_status = "Partial"
                elif a.get("employee"):
                    task_status = "OK"
                else:
                    task_status = "Unassigned"

                task_detail_rows.append({
                    "Task":         task["id"],
                    "Skill":        task["type"],
                    "Min Hrs":      task["minHours"],
                    "Assigned Hrs": round(a.get("hours", 0), 1),
                    "Assigned To":  a.get("employee") or "\u2014",
                    "Status":       task_status,
                })
            st.dataframe(pd.DataFrame(task_detail_rows), use_container_width=True, hide_index=True)


with tab_assignments:
    st.subheader("Full Task Assignment Matrix")
    matrix_rows = []
    for task in TASKS:
        a = assignments.get(task["id"], {})
        assigned_emp = a.get("employee")
        assigned_hrs = a.get("hours", 0)

        if assigned_emp and assigned_emp in emp_map:
            cost_display = f"${assigned_hrs * emp_map[assigned_emp]['rate']:,.2f}"
        else:
            cost_display = "\u2014"

        is_reimbursable = proj_map.get(task["project"], {}).get("reimbursable", False)

        if not assigned_emp:
            task_status = "Failed"
        elif a.get("partial"):
            task_status = "Partial"
        else:
            task_status = "OK"

        matrix_rows.append({
            "Task":         task["id"],
            "Project":      f"{task['project']} ({'R' if is_reimbursable else 'N'})",
            "Skill":        task["type"],
            "Min Hours":    task["minHours"],
            "Assigned Hrs": round(assigned_hrs, 1),
            "Assigned To":  assigned_emp or "\u2014",
            "Cost ($)":     cost_display,
            "Status":       task_status,
        })
    st.dataframe(pd.DataFrame(matrix_rows), use_container_width=True, hide_index=True)
